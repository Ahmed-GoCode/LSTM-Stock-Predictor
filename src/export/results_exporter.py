"""
Export functionality for stock predictor results
"""

import os
import json
import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Union
from datetime import datetime
import logging
from dataclasses import asdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
import warnings

logger = logging.getLogger(__name__)

class ResultsExporter:
    """
    Comprehensive export functionality for all stock predictor results
    """
    
    def __init__(self, output_dir: str = "outputs"):
        """
        Initialize exporter
        
        Args:
            output_dir: Directory for exported files
        """
        self.output_dir = output_dir
        self._ensure_output_directory()
        
    def _ensure_output_directory(self):
        """Ensure output directory exists"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            logger.info(f"Created output directory: {self.output_dir}")
    
    def export_predictions(self,
                         predictions: np.ndarray,
                         actual_prices: Optional[np.ndarray] = None,
                         dates: Optional[pd.DatetimeIndex] = None,
                         symbol: str = "STOCK",
                         metadata: Optional[Dict] = None,
                         format: str = "all") -> Dict[str, str]:
        """
        Export prediction results in multiple formats
        
        Args:
            predictions: Predicted prices
            actual_prices: Actual prices (optional)
            dates: Date index for predictions
            symbol: Stock symbol
            metadata: Additional metadata
            format: Export format ('csv', 'json', 'excel', 'all')
            
        Returns:
            Dictionary of exported file paths
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_filename = f"{symbol}_predictions_{timestamp}"
            
            # Prepare data
            data_dict = self._prepare_prediction_data(predictions, actual_prices, dates, symbol, metadata)
            
            exported_files = {}
            
            if format in ['csv', 'all']:
                csv_path = self._export_predictions_csv(data_dict, base_filename)
                exported_files['csv'] = csv_path
                
            if format in ['json', 'all']:
                json_path = self._export_predictions_json(data_dict, base_filename)
                exported_files['json'] = json_path
                
            if format in ['excel', 'all']:
                excel_path = self._export_predictions_excel(data_dict, base_filename)
                exported_files['excel'] = excel_path
            
            logger.info(f"Exported predictions for {symbol} in {len(exported_files)} format(s)")
            return exported_files
            
        except Exception as e:
            logger.error(f"Error exporting predictions: {e}")
            raise ValueError(f"Export failed: {e}")
    
    def _prepare_prediction_data(self,
                               predictions: np.ndarray,
                               actual_prices: Optional[np.ndarray],
                               dates: Optional[pd.DatetimeIndex],
                               symbol: str,
                               metadata: Optional[Dict]) -> Dict:
        """Prepare prediction data for export"""
        
        # Create dates if not provided
        if dates is None:
            dates = pd.date_range(start=datetime.now(), periods=len(predictions), freq='D')
        
        # Create DataFrame
        df_data = {
            'Date': dates,
            'Symbol': symbol,
            'Predicted_Price': predictions
        }
        
        if actual_prices is not None:
            df_data['Actual_Price'] = actual_prices[:len(predictions)]
            df_data['Prediction_Error'] = predictions - actual_prices[:len(predictions)]
            df_data['Absolute_Error'] = np.abs(df_data['Prediction_Error'])
            df_data['Percentage_Error'] = (df_data['Prediction_Error'] / actual_prices[:len(predictions)]) * 100
        
        df = pd.DataFrame(df_data)
        
        # Calculate summary statistics
        summary = {
            'total_predictions': len(predictions),
            'prediction_period': {
                'start_date': dates[0].isoformat(),
                'end_date': dates[-1].isoformat()
            },
            'price_statistics': {
                'min_predicted_price': float(np.min(predictions)),
                'max_predicted_price': float(np.max(predictions)),
                'mean_predicted_price': float(np.mean(predictions)),
                'std_predicted_price': float(np.std(predictions))
            }
        }
        
        if actual_prices is not None:
            error_metrics = self._calculate_prediction_metrics(predictions, actual_prices[:len(predictions)])
            summary['accuracy_metrics'] = error_metrics
        
        # Add metadata
        if metadata:
            summary['model_metadata'] = metadata
        
        summary['export_timestamp'] = datetime.now().isoformat()
        
        return {
            'dataframe': df,
            'summary': summary,
            'symbol': symbol
        }
    
    def _calculate_prediction_metrics(self, predictions: np.ndarray, actuals: np.ndarray) -> Dict:
        """Calculate prediction accuracy metrics"""
        
        errors = predictions - actuals
        
        return {
            'mae': float(np.mean(np.abs(errors))),
            'mse': float(np.mean(errors ** 2)),
            'rmse': float(np.sqrt(np.mean(errors ** 2))),
            'mape': float(np.mean(np.abs(errors / actuals)) * 100),
            'r_squared': float(1 - (np.sum(errors ** 2) / np.sum((actuals - np.mean(actuals)) ** 2))),
            'directional_accuracy': float(np.mean(np.sign(np.diff(predictions)) == np.sign(np.diff(actuals))))
        }
    
    def _export_predictions_csv(self, data_dict: Dict, base_filename: str) -> str:
        """Export predictions to CSV"""
        
        csv_path = os.path.join(self.output_dir, f"{base_filename}.csv")
        
        # Export main data
        data_dict['dataframe'].to_csv(csv_path, index=False)
        
        # Export summary as separate file
        summary_path = os.path.join(self.output_dir, f"{base_filename}_summary.csv")
        summary_df = pd.DataFrame([data_dict['summary']])
        summary_df.to_csv(summary_path, index=False)
        
        logger.info(f"Exported CSV: {csv_path}")
        return csv_path
    
    def _export_predictions_json(self, data_dict: Dict, base_filename: str) -> str:
        """Export predictions to JSON"""
        
        json_path = os.path.join(self.output_dir, f"{base_filename}.json")
        
        # Convert DataFrame to JSON-serializable format
        df = data_dict['dataframe'].copy()
        df['Date'] = df['Date'].dt.strftime('%Y-%m-%d')
        
        export_data = {
            'predictions': df.to_dict('records'),
            'summary': data_dict['summary'],
            'metadata': {
                'symbol': data_dict['symbol'],
                'export_format': 'json',
                'file_version': '1.0'
            }
        }
        
        with open(json_path, 'w') as f:
            json.dump(export_data, f, indent=2, default=str)
        
        logger.info(f"Exported JSON: {json_path}")
        return json_path
    
    def _export_predictions_excel(self, data_dict: Dict, base_filename: str) -> str:
        """Export predictions to Excel with formatting"""
        
        excel_path = os.path.join(self.output_dir, f"{base_filename}.xlsx")
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Main predictions sheet
            df = data_dict['dataframe']
            df.to_excel(writer, sheet_name='Predictions', index=False)
            
            # Summary sheet
            summary_df = pd.json_normalize(data_dict['summary'])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Formatting
            self._format_excel_workbook(writer.book, df, data_dict['summary'])
        
        logger.info(f"Exported Excel: {excel_path}")
        return excel_path
    
    def _format_excel_workbook(self, workbook: openpyxl.Workbook, df: pd.DataFrame, summary: Dict):
        """Apply formatting to Excel workbook"""
        
        try:
            # Format predictions sheet
            ws_pred = workbook['Predictions']
            
            # Header formatting
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            for cell in ws_pred[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Number formatting
            for row in ws_pred.iter_rows(min_row=2):
                for cell in row:
                    if cell.column in [3, 4]:  # Price columns
                        cell.number_format = '$#,##0.00'
                    elif cell.column in [5, 6]:  # Error columns
                        cell.number_format = '#,##0.00'
                    elif cell.column == 7:  # Percentage column
                        cell.number_format = '0.00%'
            
            # Auto-adjust column widths
            for column in ws_pred.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_pred.column_dimensions[column_letter].width = adjusted_width
            
            # Add chart if we have actual prices
            if 'Actual_Price' in df.columns:
                self._add_excel_chart(ws_pred, len(df))
                
        except Exception as e:
            logger.warning(f"Error formatting Excel file: {e}")
    
    def _add_excel_chart(self, worksheet, data_length: int):
        """Add chart to Excel worksheet"""
        
        try:
            # Create line chart
            chart = LineChart()
            chart.title = "Predicted vs Actual Prices"
            chart.style = 13
            chart.y_axis.title = 'Price ($)'
            chart.x_axis.title = 'Time Period'
            
            # Data for chart
            predicted_data = Reference(worksheet, min_col=3, min_row=1, max_row=data_length+1)
            actual_data = Reference(worksheet, min_col=4, min_row=1, max_row=data_length+1)
            dates = Reference(worksheet, min_col=1, min_row=2, max_row=data_length+1)
            
            chart.add_data(predicted_data, titles_from_data=True)
            chart.add_data(actual_data, titles_from_data=True)
            chart.set_categories(dates)
            
            # Position chart
            worksheet.add_chart(chart, "I2")
            
        except Exception as e:
            logger.warning(f"Error adding Excel chart: {e}")
    
    def export_backtest_results(self,
                              backtest_results: Dict,
                              symbol: str = "STOCK",
                              format: str = "all") -> Dict[str, str]:
        """
        Export backtesting results
        
        Args:
            backtest_results: Backtest results dictionary
            symbol: Stock symbol
            format: Export format
            
        Returns:
            Dictionary of exported file paths
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_filename = f"{symbol}_backtest_{timestamp}"
            
            exported_files = {}
            
            if format in ['csv', 'all']:
                csv_path = self._export_backtest_csv(backtest_results, base_filename)
                exported_files['csv'] = csv_path
                
            if format in ['json', 'all']:
                json_path = self._export_backtest_json(backtest_results, base_filename)
                exported_files['json'] = json_path
                
            if format in ['excel', 'all']:
                excel_path = self._export_backtest_excel(backtest_results, base_filename)
                exported_files['excel'] = excel_path
            
            logger.info(f"Exported backtest results for {symbol}")
            return exported_files
            
        except Exception as e:
            logger.error(f"Error exporting backtest results: {e}")
            raise ValueError(f"Backtest export failed: {e}")
    
    def _export_backtest_csv(self, results: Dict, base_filename: str) -> str:
        """Export backtest results to CSV"""
        
        csv_path = os.path.join(self.output_dir, f"{base_filename}.csv")
        
        # Convert results to DataFrame
        df_data = []
        for i, result in enumerate(results.get('period_results', [])):
            row = {
                'Period': i + 1,
                'Start_Date': result.get('start_date', ''),
                'End_Date': result.get('end_date', ''),
                'Total_Return': result.get('total_return', 0),
                'Sharpe_Ratio': result.get('sharpe_ratio', 0),
                'Max_Drawdown': result.get('max_drawdown', 0),
                'Win_Rate': result.get('win_rate', 0),
                'Average_Return': result.get('avg_return', 0)
            }
            df_data.append(row)
        
        df = pd.DataFrame(df_data)
        df.to_csv(csv_path, index=False)
        
        # Export summary metrics
        summary_path = os.path.join(self.output_dir, f"{base_filename}_summary.csv")
        summary_data = []
        for key, value in results.get('summary_metrics', {}).items():
            summary_data.append({'Metric': key, 'Value': value})
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_csv(summary_path, index=False)
        
        return csv_path
    
    def _export_backtest_json(self, results: Dict, base_filename: str) -> str:
        """Export backtest results to JSON"""
        
        json_path = os.path.join(self.output_dir, f"{base_filename}.json")
        
        export_data = {
            'backtest_results': results,
            'export_metadata': {
                'export_timestamp': datetime.now().isoformat(),
                'format': 'json',
                'version': '1.0'
            }
        }
        
        with open(json_path, 'w') as f:
            json.dump(export_data, f, indent=2, default=str)
        
        return json_path
    
    def _export_backtest_excel(self, results: Dict, base_filename: str) -> str:
        """Export backtest results to Excel"""
        
        excel_path = os.path.join(self.output_dir, f"{base_filename}.xlsx")
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Period results
            period_data = []
            for i, result in enumerate(results.get('period_results', [])):
                row = {
                    'Period': i + 1,
                    'Start_Date': result.get('start_date', ''),
                    'End_Date': result.get('end_date', ''),
                    'Total_Return': result.get('total_return', 0),
                    'Sharpe_Ratio': result.get('sharpe_ratio', 0),
                    'Max_Drawdown': result.get('max_drawdown', 0),
                    'Win_Rate': result.get('win_rate', 0)
                }
                period_data.append(row)
            
            period_df = pd.DataFrame(period_data)
            period_df.to_excel(writer, sheet_name='Period_Results', index=False)
            
            # Summary metrics
            summary_data = []
            for key, value in results.get('summary_metrics', {}).items():
                summary_data.append({'Metric': key, 'Value': value})
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        return excel_path
    
    def export_risk_assessment(self,
                             risk_metrics,  # RiskMetrics object
                             risk_report: Dict,
                             symbol: str = "STOCK",
                             format: str = "all") -> Dict[str, str]:
        """
        Export risk assessment results
        
        Args:
            risk_metrics: RiskMetrics object
            risk_report: Risk assessment report
            symbol: Stock symbol
            format: Export format
            
        Returns:
            Dictionary of exported file paths
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_filename = f"{symbol}_risk_assessment_{timestamp}"
            
            exported_files = {}
            
            if format in ['csv', 'all']:
                csv_path = self._export_risk_csv(risk_metrics, risk_report, base_filename)
                exported_files['csv'] = csv_path
                
            if format in ['json', 'all']:
                json_path = self._export_risk_json(risk_metrics, risk_report, base_filename)
                exported_files['json'] = json_path
                
            if format in ['excel', 'all']:
                excel_path = self._export_risk_excel(risk_metrics, risk_report, base_filename)
                exported_files['excel'] = excel_path
            
            logger.info(f"Exported risk assessment for {symbol}")
            return exported_files
            
        except Exception as e:
            logger.error(f"Error exporting risk assessment: {e}")
            raise ValueError(f"Risk assessment export failed: {e}")
    
    def _export_risk_csv(self, risk_metrics, risk_report: Dict, base_filename: str) -> str:
        """Export risk assessment to CSV"""
        
        csv_path = os.path.join(self.output_dir, f"{base_filename}.csv")
        
        # Convert risk metrics to dictionary
        if hasattr(risk_metrics, '__dict__'):
            risk_dict = asdict(risk_metrics)
        else:
            risk_dict = risk_metrics
        
        # Create DataFrame
        risk_data = []
        for key, value in risk_dict.items():
            risk_data.append({
                'Risk_Metric': key.replace('_', ' ').title(),
                'Value': value,
                'Category': self._categorize_risk_metric(key)
            })
        
        df = pd.DataFrame(risk_data)
        df.to_csv(csv_path, index=False)
        
        # Export detailed report
        report_path = os.path.join(self.output_dir, f"{base_filename}_report.csv")
        report_data = []
        
        def flatten_dict(d, parent_key='', sep='_'):
            items = []
            for k, v in d.items():
                new_key = f"{parent_key}{sep}{k}" if parent_key else k
                if isinstance(v, dict):
                    items.extend(flatten_dict(v, new_key, sep=sep).items())
                elif isinstance(v, list):
                    items.append((new_key, '; '.join(map(str, v))))
                else:
                    items.append((new_key, v))
            return dict(items)
        
        flat_report = flatten_dict(risk_report)
        for key, value in flat_report.items():
            report_data.append({'Section': key.replace('_', ' ').title(), 'Content': str(value)})
        
        report_df = pd.DataFrame(report_data)
        report_df.to_csv(report_path, index=False)
        
        return csv_path
    
    def _export_risk_json(self, risk_metrics, risk_report: Dict, base_filename: str) -> str:
        """Export risk assessment to JSON"""
        
        json_path = os.path.join(self.output_dir, f"{base_filename}.json")
        
        # Convert risk metrics to dictionary
        if hasattr(risk_metrics, '__dict__'):
            risk_dict = asdict(risk_metrics)
        else:
            risk_dict = risk_metrics
        
        export_data = {
            'risk_metrics': risk_dict,
            'risk_report': risk_report,
            'export_metadata': {
                'export_timestamp': datetime.now().isoformat(),
                'format': 'json',
                'version': '1.0'
            }
        }
        
        with open(json_path, 'w') as f:
            json.dump(export_data, f, indent=2, default=str)
        
        return json_path
    
    def _export_risk_excel(self, risk_metrics, risk_report: Dict, base_filename: str) -> str:
        """Export risk assessment to Excel"""
        
        excel_path = os.path.join(self.output_dir, f"{base_filename}.xlsx")
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Risk metrics sheet
            if hasattr(risk_metrics, '__dict__'):
                risk_dict = asdict(risk_metrics)
            else:
                risk_dict = risk_metrics
            
            risk_data = []
            for key, value in risk_dict.items():
                risk_data.append({
                    'Risk_Metric': key.replace('_', ' ').title(),
                    'Value': value,
                    'Category': self._categorize_risk_metric(key),
                    'Description': self._get_risk_metric_description(key)
                })
            
            risk_df = pd.DataFrame(risk_data)
            risk_df.to_excel(writer, sheet_name='Risk_Metrics', index=False)
            
            # Risk report sections
            for section_name, section_data in risk_report.items():
                if isinstance(section_data, dict):
                    section_df = pd.DataFrame([section_data])
                    sheet_name = section_name.replace('_', ' ').title()[:31]  # Excel sheet name limit
                    section_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return excel_path
    
    def _categorize_risk_metric(self, metric_name: str) -> str:
        """Categorize risk metric"""
        
        if 'volatility' in metric_name:
            return 'Volatility'
        elif 'var' in metric_name or 'cvar' in metric_name:
            return 'Value at Risk'
        elif 'drawdown' in metric_name:
            return 'Drawdown'
        elif metric_name in ['skewness', 'kurtosis', 'jarque_bera_stat', 'jarque_bera_p_value']:
            return 'Distribution'
        elif metric_name in ['tail_ratio', 'expected_shortfall']:
            return 'Tail Risk'
        elif metric_name in ['beta', 'correlation_with_market']:
            return 'Market Risk'
        elif 'prediction' in metric_name or 'model' in metric_name:
            return 'Model Risk'
        else:
            return 'Composite'
    
    def _get_risk_metric_description(self, metric_name: str) -> str:
        """Get description for risk metric"""
        
        descriptions = {
            'historical_volatility': 'Annualized volatility based on historical returns',
            'realized_volatility': 'Volatility calculated from squared returns',
            'garch_volatility': 'GARCH-style volatility with exponential weighting',
            'var_95': '95% Value at Risk - potential loss exceeded 5% of the time',
            'var_99': '99% Value at Risk - potential loss exceeded 1% of the time',
            'cvar_95': '95% Conditional VaR - average loss when VaR is exceeded',
            'cvar_99': '99% Conditional VaR - average loss when VaR is exceeded',
            'max_drawdown': 'Maximum peak-to-trough decline',
            'avg_drawdown': 'Average of all drawdown periods',
            'drawdown_duration': 'Average duration of drawdown periods',
            'skewness': 'Asymmetry of return distribution',
            'kurtosis': 'Tail thickness of return distribution',
            'jarque_bera_stat': 'Test statistic for normality',
            'jarque_bera_p_value': 'P-value for normality test',
            'tail_ratio': 'Ratio of extreme positive to negative returns',
            'expected_shortfall': 'Average return in worst-case scenarios',
            'beta': 'Sensitivity to market movements',
            'correlation_with_market': 'Correlation with market benchmark',
            'prediction_uncertainty': 'Uncertainty in model predictions',
            'model_confidence': 'Confidence level of model predictions',
            'overall_risk_score': 'Composite risk score (0-1 scale)',
            'risk_grade': 'Overall risk classification'
        }
        
        return descriptions.get(metric_name, 'Risk metric')
    
    def export_model_results(self,
                           model_metrics: Dict,
                           training_history: Optional[Dict] = None,
                           symbol: str = "STOCK",
                           format: str = "all") -> Dict[str, str]:
        """
        Export model training results and metrics
        
        Args:
            model_metrics: Model performance metrics
            training_history: Training history (loss, accuracy over epochs)
            symbol: Stock symbol
            format: Export format
            
        Returns:
            Dictionary of exported file paths
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_filename = f"{symbol}_model_results_{timestamp}"
            
            exported_files = {}
            
            if format in ['csv', 'all']:
                csv_path = self._export_model_csv(model_metrics, training_history, base_filename)
                exported_files['csv'] = csv_path
                
            if format in ['json', 'all']:
                json_path = self._export_model_json(model_metrics, training_history, base_filename)
                exported_files['json'] = json_path
                
            if format in ['excel', 'all']:
                excel_path = self._export_model_excel(model_metrics, training_history, base_filename)
                exported_files['excel'] = excel_path
            
            logger.info(f"Exported model results for {symbol}")
            return exported_files
            
        except Exception as e:
            logger.error(f"Error exporting model results: {e}")
            raise ValueError(f"Model results export failed: {e}")
    
    def _export_model_csv(self, metrics: Dict, history: Optional[Dict], base_filename: str) -> str:
        """Export model results to CSV"""
        
        csv_path = os.path.join(self.output_dir, f"{base_filename}.csv")
        
        # Model metrics
        metrics_data = []
        for key, value in metrics.items():
            metrics_data.append({'Metric': key, 'Value': value})
        
        metrics_df = pd.DataFrame(metrics_data)
        metrics_df.to_csv(csv_path, index=False)
        
        # Training history
        if history:
            history_path = os.path.join(self.output_dir, f"{base_filename}_history.csv")
            history_df = pd.DataFrame(history)
            history_df.to_csv(history_path, index=False)
        
        return csv_path
    
    def _export_model_json(self, metrics: Dict, history: Optional[Dict], base_filename: str) -> str:
        """Export model results to JSON"""
        
        json_path = os.path.join(self.output_dir, f"{base_filename}.json")
        
        export_data = {
            'model_metrics': metrics,
            'training_history': history,
            'export_metadata': {
                'export_timestamp': datetime.now().isoformat(),
                'format': 'json',
                'version': '1.0'
            }
        }
        
        with open(json_path, 'w') as f:
            json.dump(export_data, f, indent=2, default=str)
        
        return json_path
    
    def _export_model_excel(self, metrics: Dict, history: Optional[Dict], base_filename: str) -> str:
        """Export model results to Excel"""
        
        excel_path = os.path.join(self.output_dir, f"{base_filename}.xlsx")
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Model metrics
            metrics_data = []
            for key, value in metrics.items():
                metrics_data.append({
                    'Metric': key.replace('_', ' ').title(),
                    'Value': value,
                    'Description': self._get_metric_description(key)
                })
            
            metrics_df = pd.DataFrame(metrics_data)
            metrics_df.to_excel(writer, sheet_name='Model_Metrics', index=False)
            
            # Training history
            if history:
                history_df = pd.DataFrame(history)
                history_df.to_excel(writer, sheet_name='Training_History', index=False)
        
        return excel_path
    
    def _get_metric_description(self, metric_name: str) -> str:
        """Get description for model metric"""
        
        descriptions = {
            'mse': 'Mean Squared Error - average of squared prediction errors',
            'rmse': 'Root Mean Squared Error - square root of MSE',
            'mae': 'Mean Absolute Error - average of absolute prediction errors',
            'mape': 'Mean Absolute Percentage Error - average percentage error',
            'r2_score': 'R-squared - proportion of variance explained by model',
            'directional_accuracy': 'Percentage of correct directional predictions',
            'training_time': 'Time taken to train the model',
            'total_parameters': 'Total number of model parameters',
            'final_loss': 'Final training loss value',
            'final_val_loss': 'Final validation loss value'
        }
        
        return descriptions.get(metric_name, 'Model performance metric')
    
    def create_comprehensive_report(self,
                                  symbol: str,
                                  predictions_data: Dict,
                                  backtest_results: Dict,
                                  risk_assessment: Dict,
                                  model_results: Dict,
                                  format: str = "excel") -> str:
        """
        Create comprehensive report combining all results
        
        Args:
            symbol: Stock symbol
            predictions_data: Prediction results
            backtest_results: Backtesting results
            risk_assessment: Risk assessment results
            model_results: Model training results
            format: Export format
            
        Returns:
            Path to comprehensive report file
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{symbol}_comprehensive_report_{timestamp}"
            
            if format == "excel":
                return self._create_excel_comprehensive_report(
                    filename, symbol, predictions_data, backtest_results, 
                    risk_assessment, model_results
                )
            elif format == "json":
                return self._create_json_comprehensive_report(
                    filename, symbol, predictions_data, backtest_results,
                    risk_assessment, model_results
                )
            else:
                raise ValueError(f"Unsupported format for comprehensive report: {format}")
                
        except Exception as e:
            logger.error(f"Error creating comprehensive report: {e}")
            raise ValueError(f"Comprehensive report creation failed: {e}")
    
    def _create_excel_comprehensive_report(self,
                                         filename: str,
                                         symbol: str,
                                         predictions_data: Dict,
                                         backtest_results: Dict,
                                         risk_assessment: Dict,
                                         model_results: Dict) -> str:
        """Create comprehensive Excel report"""
        
        excel_path = os.path.join(self.output_dir, f"{filename}.xlsx")
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Executive Summary
            summary_data = {
                'Symbol': [symbol],
                'Report_Date': [datetime.now().strftime('%Y-%m-%d')],
                'Model_Type': ['LSTM Neural Network'],
                'Overall_Risk_Grade': [risk_assessment.get('risk_metrics', {}).get('risk_grade', 'Unknown')],
                'Backtest_Sharpe_Ratio': [backtest_results.get('summary_metrics', {}).get('sharpe_ratio', 0)],
                'Model_Accuracy_R2': [model_results.get('r2_score', 0)],
                'Prediction_Confidence': [risk_assessment.get('risk_metrics', {}).get('model_confidence', 0)]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Executive_Summary', index=False)
            
            # Add other sheets using existing methods
            # (This would integrate with the individual export methods)
        
        logger.info(f"Created comprehensive Excel report: {excel_path}")
        return excel_path
    
    def _create_json_comprehensive_report(self,
                                        filename: str,
                                        symbol: str,
                                        predictions_data: Dict,
                                        backtest_results: Dict,
                                        risk_assessment: Dict,
                                        model_results: Dict) -> str:
        """Create comprehensive JSON report"""
        
        json_path = os.path.join(self.output_dir, f"{filename}.json")
        
        comprehensive_report = {
            'metadata': {
                'symbol': symbol,
                'report_date': datetime.now().isoformat(),
                'model_type': 'LSTM Neural Network',
                'report_version': '1.0'
            },
            'executive_summary': {
                'overall_risk_grade': risk_assessment.get('risk_metrics', {}).get('risk_grade', 'Unknown'),
                'model_confidence': risk_assessment.get('risk_metrics', {}).get('model_confidence', 0),
                'backtest_performance': backtest_results.get('summary_metrics', {}),
                'key_metrics': {
                    'prediction_accuracy': model_results.get('r2_score', 0),
                    'volatility': risk_assessment.get('risk_metrics', {}).get('historical_volatility', 0),
                    'max_drawdown': risk_assessment.get('risk_metrics', {}).get('max_drawdown', 0)
                }
            },
            'detailed_results': {
                'predictions': predictions_data,
                'backtesting': backtest_results,
                'risk_assessment': risk_assessment,
                'model_performance': model_results
            }
        }
        
        with open(json_path, 'w') as f:
            json.dump(comprehensive_report, f, indent=2, default=str)
        
        logger.info(f"Created comprehensive JSON report: {json_path}")
        return json_path